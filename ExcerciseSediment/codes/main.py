# This is main.py
import pandas as pd
import os
from grains import GrainReader

from hec import HecSet
from mpm import *

# Importiere die Funktion für die Formatierung
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill





def get_char_grain_size(file_name=str, D_char=str):
    grain_info = GrainReader(file_name)
    return grain_info.size_classes["size"][D_char]


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import os
import win32com.client as win32

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def format_excel_headers(file_name, sheet_name):
    """
    Add a background color to the headers of an output table in an Excel file.
    :param file_name: Path to the Excel file.
    :param sheet_name: Name of the sheet to format.
    """
    # Load the workbook and select the desired sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define a fill style for the header using RGB for red
    header_fill = PatternFill(start_color="D5006D", end_color="D5006D", fill_type="solid")

    # Apply the style to the header row (assuming headers are in the first row)
    for cell in ws[1]:
        print(f"Formatting cell: {cell.value}")  # Confirm which cells are being formatted
        cell.fill = header_fill

    # Save the workbook
    wb.save(file_name)
    print(f"Headers in {sheet_name} have been formatted.")




def save_excel_as_pdf(excel_file, pdf_file):
    """
    Save the Excel file as a PDF.
    :param excel_file: Path to the Excel file.
    :param pdf_file: Path to save the PDF file.
    """
    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False  # Excel window remains hidden during operation
    wb = excel.Workbooks.Open(excel_file)

    # Export as PDF
    wb.ExportAsFixedFormat(0, pdf_file)
    wb.Close(False)
    excel.Quit()
    print(f"Excel file saved as PDF: {pdf_file}")




import pandas as pd
import matplotlib.pyplot as plt

def plot_bedload_vs_discharge(results_file, profiles):
    """
    Plot dimensional bedload transport against discharge for selected profiles.
    :param results_file: Path to the mpm_results file (Excel format).
    :param profiles: List of profile identifiers to plot.
    """
    # Load the results from Excel file
    df = pd.read_excel(results_file)

    # Filter data for the chosen profiles
    selected_profiles = df[df['Scenario'].isin(profiles)]  # Hier 'Scenario' oder das entsprechende Profilfeld anpassen

    # Plot each profile
    plt.figure(figsize=(10, 6))
    for profile in profiles:
        profile_data = selected_profiles[selected_profiles['Scenario'] == profile]
        plt.plot(
            profile_data['Q (m3/s)'],  # Discharge (x-axis)
            profile_data['Qb (kg/s)'],  # Bedload Transport (y-axis)
            label=f'Profile {profile}'
        )

    # Customize the plot
    plt.title('Dimensional Bedload Transport vs. Discharge')
    plt.xlabel('Discharge (m³/s)')
    plt.ylabel('Bedload Transport (kg/s)')
    plt.legend()
    plt.grid(True)
    plt.show()


# Deine bestehende Funktion für das Berechnen des MPM
def calculate_mpm(hec_df, D_char):
    mpm_dict = {
        "River Sta": [],
        "Scenario": [],
        "Q (m3/s)": [],
        "Phi (-)": [],
        "Qb (kg/s)": []
    }

    Froude = hec_df["Froude # Chl"]
    h = hec_df["Hydr Depth"]
    Q = hec_df["Q Total"]
    Rh = hec_df["Hydr Radius"]
    Se = hec_df["E.G. Slope"]
    u = hec_df["Vel Chnl"]

    for i, sta in enumerate(list(hec_df["River Sta"])):
        if not str(sta).lower() == "nan":
            logging.info("PROCESSING PROFILE {0} FOR SCENARIO {1}".format(str(hec_df["River Sta"][i]),
                                                                          str(hec_df["Profile"][i])))
            mpm_dict["River Sta"].append(hec_df["River Sta"][i])
            mpm_dict["Scenario"].append(hec_df["Profile"][i])
            section_mpm = MPM(grain_size=D_char,
                              Froude=Froude[i],
                              water_depth=h[i],
                              velocity=u[i],
                              Q=Q[i],
                              hydraulic_radius=Rh[i],
                              slope=Se[i])
            mpm_dict["Q (m3/s)"].append(Q[i])
            mpm_dict["Phi (-)"].append(section_mpm.phi)
            b = hec_df["Flow Area"][i] / h[i]
            mpm_dict["Qb (kg/s)"].append(section_mpm.add_dimensions(b))
    return pd.DataFrame(mpm_dict)


# Hauptfunktion
@log_actions
def main():
    D_char = get_char_grain_size(file_name=os.path.abspath("..") + "\\grains.csv", D_char="D84")
    logging.info("This is the characteristic grain size: %0.4f m." % (D_char))

    hec_file = os.path.abspath("..") + "\\HEC-RAS\\output.xlsx"
    hec = HecSet(hec_file)
    logging.info(hec.hec_data.head())

    mpm_results = calculate_mpm(hec.hec_data, D_char)
    mpm_results.to_excel(os.path.abspath("..") + "\\bed_load_mpm.xlsx", index=False)

    # Format headers in the Excel file
    format_excel_headers(os.path.abspath("..") + "\\bed_load_mpm.xlsx", "Sheet1")

    # Save the Excel file as PDF
    #pdf_file = os.path.abspath("..") + "\\bed_load_mpm.pdf"
    #save_excel_as_pdf(os.path.abspath("..") + "\\bed_load_mpm.xlsx", pdf_file)

    #results_file = os.path.abspath("..") + "\\bed_load_mpm.xlsx"  # Path to your Excel file
    #profiles_to_plot = [1, 2, 3]  # Profile IDs you want to plot
    #plot_bedload_vs_discharge(results_file, profiles_to_plot)


if __name__ == '__main__':
    main()
